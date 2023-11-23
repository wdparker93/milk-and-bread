import openpyxl
import os
import mysql.connector

#Executes array of insertion queries on the database
def insertToDatabase(insertQueryArr):
    dbConnection = mysql.connector.connect(user="root", password="password", host="localhost", database="milk_and_bread_schema")
    if dbConnection.is_connected():
        print('Connection to MySQL DB Successful')
    else:
        print('ERROR: Did not connect to MySQL DB')
        exit()
    crsr = dbConnection.cursor()
    for i in range(len(insertQueryArr)):
        #print(insertQueryArr[i])
        crsr.execute(insertQueryArr[i])
        dbConnection.commit()
    dbConnection.close()

#Takes an array of parameters for sql inserts and builds statements with it
def generateInsertStatements(paramArray):
    returnStatementArray = []
    for i in range(len(paramArray)):
        sqlInsertStatement = "INSERT INTO grocery_sales_data(date, state, category, dollars_sold, units_sold) VALUES ("
        sqlInsertStatement += "'" + paramArray[i][0] + "', "    #date
        sqlInsertStatement += "'" + paramArray[i][1] + "', "    #state
        sqlInsertStatement += "'" + paramArray[i][2] + "', "    #category
        sqlInsertStatement += str(paramArray[i][3]) + ", "      #dollars_sold
        sqlInsertStatement += str(paramArray[i][4])             #units_sold
        sqlInsertStatement += ");"
        returnStatementArray.append(sqlInsertStatement)
    return returnStatementArray

#Processes the Excel file into an array of parameters to be inserted into the database
def processExcelFile():
    curDir = os.getcwd()
    dataDir = curDir + '\\raw_data\\StateAndCategory.xlsx'
    print('Loading workbook at ' + dataDir + '. Please wait...')

    groceryDataWB = openpyxl.load_workbook(dataDir)
    print(dataDir + ' loaded successfully.')
    groceryDataSheet = groceryDataWB.active

    #Column defintions
    dateCol = 1
    stateCol = 2
    categoryCol = 3
    dollarsCol = 4
    unitsCol = 5

    #Initialze to first row of data
    entryRow = 3

    #Array that will hold all sets of insert parameters
    parentArray = []
    #Array that will hold an individual set of insert parameters
    childArray = []

    entryCell = groceryDataSheet.cell(row = entryRow, column = dateCol)
    while entryCell.value != 'NA = data are not available':
        childArray = []
        if entryCell.value != "" and entryCell.value != None:
            childArray.append((groceryDataSheet.cell(row = entryRow, column = dateCol)).value)  #date
            childArray.append((groceryDataSheet.cell(row = entryRow, column = stateCol)).value) #state
            childArray.append((groceryDataSheet.cell(row = entryRow, column = categoryCol)).value)  #category
            dollarsSold = (groceryDataSheet.cell(row = entryRow, column = dollarsCol)).value
            if (dollarsSold == 'NA'):
                dollarsSold = 0
            childArray.append(dollarsSold)  #dollars_sold
            unitsSold = (groceryDataSheet.cell(row = entryRow, column = unitsCol)).value
            if (unitsSold == 'NA'):
                unitsSold = 0
            childArray.append(unitsSold)  #units_sold
            #print(childArray)
            parentArray.append(childArray)
        entryRow += 1
        entryCell = groceryDataSheet.cell(row = entryRow, column = dateCol)
    return parentArray


sqlInsertParamArray = processExcelFile()
sqlInsertStatementArray = generateInsertStatements(sqlInsertParamArray)
insertToDatabase(sqlInsertStatementArray)